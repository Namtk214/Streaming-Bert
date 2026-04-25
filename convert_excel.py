"""
Convert Excel test data → JSON format cho Early-Exit with Noisy-OR Loss.

Input:  Excel file với 2 sheets:
          Sheet1   — scam conversations
          no_scam  — harmless conversations
        Mỗi row: (index, full_conversation_string)
        Turns cách nhau bằng '\n'.

Output: JSON file cùng format với processed data từ prepare_data.py:
  [
    {
      "dialogue_id":         "scam-1",
      "dialogue_label":      1,
      "dialogue_label_name": "scam",
      "num_turns":           5,
      "turns": [
        {
          "turn_id":        1,
          "role":           "người gọi",
          "text":           "...",
          "text_segmented": "..."
        }, ...
      ]
    }, ...
  ]

Usage:
    python convert_excel.py
    python convert_excel.py --excel ../../"Tổng hợp...xlsx" --out data/excel_test.json
"""

import argparse
import json
import os
import sys
import unicodedata
import re

_streaming_dir = os.path.dirname(os.path.abspath(__file__))
if _streaming_dir not in sys.path:
    sys.path.insert(0, _streaming_dir)

from config import EarlyExitConfig, LABEL_MAP, VNCORENLP_CACHE

cfg = EarlyExitConfig()

_DEFAULT_EXCEL = os.path.join(
    os.path.dirname(_streaming_dir),
    "Tổng hợp kịch bản test AI on devices_result_v2.xlsx",
)
_DEFAULT_OUT = os.path.join(cfg.data_dir, "excel_test.json")

SHEET_LABEL = {"Sheet1": "scam", "no_scam": "harmless"}

# Speaker roles: alternate between caller and receiver
ROLE_MAP = {0: "người gọi", 1: "người nghe"}


# ── Text utilities ───────────────────────────────────────────────

def clean_text(text: str) -> str:
    text = unicodedata.normalize("NFC", text)
    text = re.sub(r"[\x00-\x09\x0b-\x0c\x0e-\x1f\x7f]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


# ── Word segmenter (reuses VNCORENLP_CACHE) ─────────────────────

class WordSegmenter:
    def __init__(self, vncorenlp_dir: str):
        import py_vncorenlp
        abs_dir = os.path.abspath(vncorenlp_dir)
        if abs_dir in VNCORENLP_CACHE:
            self.segmenter = VNCORENLP_CACHE[abs_dir]
            print("  VnCoreNLP reused from cache")
            return
        jar = os.path.join(abs_dir, "VnCoreNLP-1.2.jar")
        if not os.path.exists(jar):
            print(f"  Downloading VnCoreNLP into {abs_dir} ...")
            os.makedirs(abs_dir, exist_ok=True)
            py_vncorenlp.download_model(save_dir=abs_dir)
        self.segmenter = py_vncorenlp.VnCoreNLP(annotators=["wseg"], save_dir=abs_dir)
        VNCORENLP_CACHE[abs_dir] = self.segmenter
        print("  VnCoreNLP loaded OK")

    def segment(self, text: str) -> str:
        result = self.segmenter.word_segment(text)
        return " ".join(result) if isinstance(result, list) else result


# ── Excel reader ─────────────────────────────────────────────────

def load_excel(path: str) -> list:
    """Return list of (sheet_label, row_idx, raw_turns_list)."""
    try:
        import openpyxl
    except ImportError:
        print("  [ERROR] openpyxl chưa cài. Chạy: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = []
    for sheet_name, label in SHEET_LABEL.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Không tìm thấy sheet '{sheet_name}', bỏ qua.")
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, values_only=True):
            idx, conv_text = row[0], row[1]
            if not conv_text or not str(conv_text).strip():
                continue
            raw_turns = [t.strip() for t in str(conv_text).strip().split("\n") if t.strip()]
            if raw_turns:
                records.append((label, idx, raw_turns))
    return records


# ── Convert one conversation ─────────────────────────────────────

def convert_conversation(label: str, idx, raw_turns: list, segmenter: WordSegmenter) -> dict:
    """Convert 1 dialogue sang format Noisy-OR processed data."""
    dialogue_label = LABEL_MAP[label]  # "harmless" → 0, "scam" → 1

    turns = []
    for i, raw_text in enumerate(raw_turns):
        text = clean_text(raw_text)
        # Alternating speaker: even turn index → người gọi, odd → người nghe
        role = ROLE_MAP[i % 2]
        turns.append({
            "turn_id":        i + 1,
            "role":           role,
            "text":           text,
            "text_segmented": segmenter.segment(text),
        })

    return {
        "dialogue_id":         f"{label}-{idx}",
        "dialogue_label":      dialogue_label,
        "dialogue_label_name": label,
        "num_turns":           len(turns),
        "turns":               turns,
    }


# ── Main ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Convert Excel test data to JSON (Noisy-OR format).")
    parser.add_argument("--excel",      default=_DEFAULT_EXCEL)
    parser.add_argument("--out",        default=_DEFAULT_OUT)
    parser.add_argument("--vncorenlp",  default=cfg.vncorenlp_dir)
    args = parser.parse_args()

    print("=" * 60)
    print("CONVERT EXCEL → NOISY-OR JSON")
    print("=" * 60)
    print(f"  Excel:      {args.excel}")
    print(f"  Output:     {args.out}")
    print(f"  VnCoreNLP:  {args.vncorenlp}")

    if not os.path.exists(args.excel):
        print(f"\n  [ERROR] Không tìm thấy: {args.excel}")
        sys.exit(1)

    print("\nLoading Excel...")
    records = load_excel(args.excel)
    scam_n = sum(1 for r in records if r[0] == "scam")
    harm_n = sum(1 for r in records if r[0] == "harmless")
    print(f"  {len(records)} conversations (scam={scam_n}, harmless={harm_n})")

    print("\nInitialising word segmenter...")
    segmenter = WordSegmenter(args.vncorenlp)

    print("\nConverting...")
    converted = []
    for i, (label, idx, raw_turns) in enumerate(records, 1):
        converted.append(convert_conversation(label, idx, raw_turns, segmenter))
        if i % 100 == 0:
            print(f"  {i}/{len(records)} done")

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(converted, f, ensure_ascii=False, indent=2)

    print(f"\n  Saved {len(converted)} conversations → {args.out}")

    # Sample
    s = converted[0]
    print(f"\n-- Sample ({s['dialogue_id']}) --")
    print(f"   Label: {s['dialogue_label_name']} ({s['dialogue_label']})")
    print(f"   Turns: {s['num_turns']}")
    for t in s["turns"][:3]:
        print(f"   T{t['turn_id']} ({t['role']}): {t['text'][:60]}...")
    print("\nDone!")


if __name__ == "__main__":
    main()
